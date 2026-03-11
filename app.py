# ============================================================
# 🚢 app.py — DG Cargo Guardian 主介面（完整版）
# ============================================================

import json
import io
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
from itertools import combinations

import streamlit as st
from ems_engine       import query_ems, format_ems_report
from ai_analyzer      import analyze_incident, ask_dg_question, check_segregation, INCIDENT_SOP_MAP
from manifest_parser  import (
    parse_manifest_excel, parse_manifest_csv,
    parse_asc_file,
    get_manifest_summary, generate_sample_template,
)
from bay_plan_engine  import (
    build_bay_plan, get_bay_dimensions,
    get_cell_display, get_plan_statistics,
    get_row_label, get_tier_label,
)
from fire_classifier  import classify_fire_category, get_color_legend
from llm_client       import get_llm_response
from ai_analyzer      import SYSTEM_PROMPT


# ══════════════════════════════════════════════════════════════
# ── 積載隔離輔助函數
# ══════════════════════════════════════════════════════════════

def _validate_position(pos: str) -> bool:
    """驗證 BBRRTT 格式：6位純數字"""
    return len(pos) == 6 and pos.isdigit()


def _format_position(pos: str) -> str:
    """將 BBRRTT 轉為可讀格式"""
    if not _validate_position(pos):
        return pos

    bb     = pos[0:2]
    rr     = pos[2:4]
    tt     = pos[4:6]
    rr_int = int(rr)
    tt_int = int(tt)

    # Row 描述：00=中心線，奇數=左舷，偶數=右舷
    if rr_int == 0:
        row_desc = "中心線"
    elif rr_int % 2 == 1:
        row_desc = f"左舷第{(rr_int + 1) // 2}列"
    else:
        row_desc = f"右舷第{rr_int // 2}列"

    # Tier 描述
    if tt_int < 80:
        tier_desc = f"艙內第{(tt_int - 2) // 2 + 1}層"
    else:
        tier_desc = f"甲板上第{(tt_int - 82) // 2 + 1}層"

    return f"Bay{bb} Row{rr}({row_desc}) Tier{tt}({tier_desc})"


def _calc_distance(pos_a: str, pos_b: str) -> str:
    """計算兩個貨櫃位置的大略距離描述"""
    if not (_validate_position(pos_a) and _validate_position(pos_b)):
        return "未知"

    bay_a  = int(pos_a[0:2])
    row_a  = int(pos_a[2:4])
    tier_a = int(pos_a[4:6])
    bay_b  = int(pos_b[0:2])
    row_b  = int(pos_b[2:4])
    tier_b = int(pos_b[4:6])

    bay_dist  = abs(bay_a - bay_b) * 6.0
    row_dist  = abs(row_a - row_b) * 2.4
    tier_dist = abs(tier_a - tier_b) * 2.6
    total     = (bay_dist**2 + row_dist**2 + tier_dist**2) ** 0.5

    if total == 0:
        return "同一位置"
    elif total < 3:
        return f"約 {total:.1f}m（緊鄰）"
    elif total < 12:
        return f"約 {total:.1f}m（近距）"
    else:
        return f"約 {total:.1f}m"


def _render_position_map(cargos: list):
    """用 Streamlit 繪製簡易 Bay/Row 平面示意圖"""
    positions = []
    for c in cargos:
        pos = c["position"]
        if _validate_position(pos):
            positions.append({
                "label":   c["label"],
                "un":      c["un"],
                "bay":     int(pos[0:2]),
                "row":     int(pos[2:4]),
                "tier":    int(pos[4:6]),
                "on_deck": int(pos[4:6]) >= 80,
                "class":   c["data"]["hazard_class"],
            })

    if not positions:
        return

    all_bays = sorted(set(p["bay"] for p in positions))
    all_rows = sorted(set(p["row"] for p in positions))

    grid = {}
    for bay in all_bays:
        col_data = {}
        for row in all_rows:
            items = [p for p in positions if p["bay"] == bay and p["row"] == row]
            col_data[f"Row {row:02d}"] = (
                " / ".join(f"{p['label']}(UN{p['un']})" for p in items)
                if items else "—"
            )
        grid[f"Bay {bay:02d}"] = col_data

    df = pd.DataFrame(grid).T
    st.dataframe(df, use_container_width=True)

    colors = ["🔴", "🔵", "🟢", "🟡", "🟠", "🟣", "⚫", "⚪", "🟤", "🔶"]
    cols   = st.columns(len(cargos))
    for i, cargo in enumerate(cargos):
        if _validate_position(cargo["position"]):
            cols[i].caption(
                f"{colors[i % len(colors)]} {cargo['label']} | "
                f"UN{cargo['un']} | {_format_position(cargo['position'])}"
            )


def _generate_segregation_report(cargos: list, results: list, violation_count: int) -> str:
    """產生純文字隔離檢查報告"""
    lines = [
        "=" * 60,
        "  DG CARGO GUARDIAN — 積載隔離檢查報告",
        f"  產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60, "",
        "【貨物清單】",
    ]
    for c in cargos:
        lines.append(
            f"  {c['label']:6s} | UN{c['un']:4s} | "
            f"{c['data']['proper_shipping_name'][:35]:35s} | "
            f"Class {c['data']['hazard_class']:4s} | "
            f"位置：{_format_position(c['position'])}"
        )
    lines += [
        "",
        f"【檢查結果摘要】共 {len(results)} 組配對，{violation_count} 項違規",
        "",
    ]
    for res in results:
        is_v = any(
            kw in res["result"]
            for kw in ["違規", "違反", "不得", "禁止", "VIOLATION", "❌"]
        )
        lines += [
            "-" * 60,
            f"{'🚨 [違規]' if is_v else '✅ [合規]'}  "
            f"{res['label_a']}(UN{res['un_a']} @ {res['pos_a']})  ×  "
            f"{res['label_b']}(UN{res['un_b']} @ {res['pos_b']})",
            f"距離：{_calc_distance(res['pos_a'], res['pos_b'])}",
            "",
            res["result"],
            "",
        ]
    lines += [
        "=" * 60,
        "⚠️  本報告僅供參考，實際操作請依 IMDG Code 官方規定",
        "=" * 60,
    ]
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════
# ── 頁面設定
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="DG Cargo Guardian",
    page_icon="⚠️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 自訂 CSS ─────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 2rem;
        font-weight: 700;
        color: #FF4B4B;
        text-align: center;
        padding: 1rem 0 0.2rem 0;
    }
    .sub-title {
        font-size: 1rem;
        color: #888;
        text-align: center;
        margin-bottom: 2rem;
    }
    .info-card {
        background: #2A1F2E;
        border-left: 4px solid #FF4B4B;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0;
        color: #F0F0F0;
    }
    .info-card-blue {
        background: #1A2535;
        border-left: 4px solid #4B9EFF;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0;
        color: #F0F0F0;
    }
    .info-card-green {
        background: #1A2B25;
        border-left: 4px solid #4BFF9E;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0;
        color: #F0F0F0;
    }
    .ems-badge {
        display: inline-block;
        background: #FF4B4B;
        color: white;
        font-size: 1.1rem;
        font-weight: 700;
        padding: 0.3rem 0.8rem;
        border-radius: 6px;
        margin-right: 0.5rem;
        letter-spacing: 1px;
    }
    .ems-badge-blue {
        display: inline-block;
        background: #4B9EFF;
        color: white;
        font-size: 1.1rem;
        font-weight: 700;
        padding: 0.3rem 0.8rem;
        border-radius: 6px;
        margin-right: 0.5rem;
        letter-spacing: 1px;
    }
    .ai-response-wrapper {
        background: #F8F9FA;
        border: 1px solid #DEE2E6;
        border-radius: 10px;
        padding: 1.5rem 1.8rem;
        margin-top: 0.5rem;
    }
    .warning-banner {
        background: #3D1F00;
        border: 1px solid #FF8C00;
        border-radius: 8px;
        padding: 0.7rem 1rem;
        color: #FFB347;
        font-size: 0.85rem;
        text-align: center;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)


# ── Session State 初始化 ─────────────────────────────────────
for key, default in {
    "chat_history":   [],
    "last_un":        "",
    "cargo_list":     [
        {"un": "", "position": "", "label": "貨物 1"},
        {"un": "", "position": "", "label": "貨物 2"},
    ],
    "dg_cargo_list":  [],
    "dg_bay_plan":    {},
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════════════════════
# ── 側邊欄
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.image("https://img.icons8.com/color/96/hazmat.png", width=80)
    st.markdown("## ⚠️ DG Cargo Guardian")
    st.markdown("*海運危險品應急處置系統*")
    st.divider()

    page = st.radio(
        "📋 功能選單",
        options=[
            "🔍 EMS 快速查詢",
            "🤖 AI 事故分析",
            "🔄 積載隔離檢查",
            "🗺️ DG Bay Plan",
            "💬 自由問答",
        ],
        label_visibility="collapsed"
    )

    st.divider()

    st.markdown("#### 🔖 常用 UN 號碼")
    quick_uns = {
        "UN1203 — 汽油":  "1203",
        "UN1017 — 氯氣":  "1017",
        "UN1789 — 鹽酸":  "1789",
        "UN3480 — 鋰電池":"3480",
        "UN1072 — 氧氣":  "1072",
    }
    for label, un in quick_uns.items():
        if st.button(label, use_container_width=True):
            st.session_state.last_un = un

    st.divider()
    st.markdown(
        "<div style='font-size:0.75rem; color:#666; text-align:center;'>"
        "⚠️ 本系統僅供參考<br>實際操作請依官方 IMDG Code"
        "</div>",
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════════════
# 頁面 1：EMS 快速查詢
# ══════════════════════════════════════════════════════════════
if page == "🔍 EMS 快速查詢":

    st.markdown('<div class="main-title">🔍 EMS 快速查詢</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">輸入 UN 號碼，即時取得 IMDG 應急程序資料</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        un_input = st.text_input(
            "UN 號碼",
            value=st.session_state.last_un,
            placeholder="例如：1203",
            label_visibility="collapsed"
        )
    with col2:
        search_btn = st.button("🔍 查詢", use_container_width=True, type="primary")

    if search_btn and un_input:
        st.session_state.last_un = un_input.strip()
        data = query_ems(un_input.strip())

        if not data["found"]:
            st.error(f"❌ {data['message']}")
        else:
            st.markdown("---")
            col_a, col_b, col_c, col_d = st.columns(4)
            col_a.metric("UN 號碼",   data["un_number"])
            col_b.metric("危險品類別", f"Class {data['hazard_class']}")
            col_c.metric("包裝等級",   data["packing_group"] or "N/A")
            col_d.metric("MFAG",       data["mfag"] or "N/A")

            st.markdown(f"### 📦 {data['proper_shipping_name']}")
            st.caption(data["description"])

            st.markdown("#### 🚨 EMS 應急程序代碼")
            ems      = data["ems"]
            col_e, col_f = st.columns(2)
            with col_e:
                st.markdown(
                    f'<div class="info-card">'
                    f'<span class="ems-badge">🔥 {ems["fire_code"]}</span>'
                    f'<br><br><b>{ems["fire_description"]}</b>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            with col_f:
                st.markdown(
                    f'<div class="info-card-blue">'
                    f'<span class="ems-badge-blue">💧 {ems["spillage_code"]}</span>'
                    f'<br><br><b>{ems["spillage_description"]}</b>'
                    f'</div>',
                    unsafe_allow_html=True
                )

            st.markdown("#### 🛡️ 應急處置指引")
            ea = data.get("emergency_action", {})
            tab1, tab2, tab3 = st.tabs(["🔥 火災處置", "💧 洩漏處置", "🏥 急救處置"])
            with tab1:
                st.info(ea.get("fire",      "無資料"))
            with tab2:
                st.info(ea.get("spillage",  "無資料"))
            with tab3:
                st.info(ea.get("first_aid", "無資料"))

            st.markdown("#### 📍 積載與隔離")
            col_g, col_h = st.columns(2)
            with col_g:
                st.markdown(
                    f'<div class="info-card-green">'
                    f'<b>積載類別</b><br>{data["stowage"] or "N/A"}'
                    f'</div>',
                    unsafe_allow_html=True
                )
            with col_h:
                sp = "、".join(data["special_provisions"]) if data["special_provisions"] else "無"
                st.markdown(
                    f'<div class="info-card-green">'
                    f'<b>特殊規定</b><br>{sp}'
                    f'</div>',
                    unsafe_allow_html=True
                )

            with st.expander("📄 查看完整原始報告"):
                st.code(format_ems_report(data), language="text")


# ══════════════════════════════════════════════════════════════
# 頁面 2：AI 事故分析
# ══════════════════════════════════════════════════════════════
elif page == "🤖 AI 事故分析":

    st.markdown('<div class="main-title">🤖 AI 事故分析</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">描述事故情境，AI 根據 IMDG 資料給出應急建議</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="warning-banner">'
        '⚠️ AI 建議僅供參考，緊急情況請立即聯繫 CHEMTREC (+1-703-527-3887) 或當地應急機構'
        '</div>',
        unsafe_allow_html=True
    )

    col1, col2 = st.columns([1, 1])
    with col1:
        un_input = st.text_input(
            "UN 號碼",
            value=st.session_state.last_un,
            placeholder="例如：1203"
        )
    with col2:
        incident_type = st.selectbox(
            "事故類型",
            options=[
                "deck_container_fire",
                "hold_container_fire",
                "engine_room_fire",
                "cargo_leakage",
                "container_overboard",
                "fire",
                "spillage",
                "first_aid",
                "general",
            ],
            format_func=lambda x: {
                "deck_container_fire": "🔥 甲板貨櫃失火（WHL 3-3）",
                "hold_container_fire": "🔥 貨艙貨櫃失火（WHL 3-4）",
                "engine_room_fire":    "🔥 機艙失火（WHL 1-5）",
                "cargo_leakage":       "💧 貨櫃洩漏 氣體/液體（WHL 3-5）",
                "container_overboard": "📦 貨櫃落海/傾倒/位移（WHL 3-2）",
                "fire":                "🔥 火災事故（一般）",
                "spillage":            "💧 洩漏事故（一般）",
                "first_aid":           "🏥 人員傷亡急救（MFAG）",
                "general":             "📋 一般查詢",
            }[x],
            label_visibility="collapsed"
        )

    # SOP 參考標籤
    sop_badges = {
        "deck_container_fire": ("3-3", "#dc2626"),
        "hold_container_fire": ("3-4", "#b45309"),
        "engine_room_fire":    ("1-5", "#7c3aed"),
        "cargo_leakage":       ("3-5", "#0369a1"),
        "container_overboard": ("3-2", "#047857"),
        "fire":                ("IMDG", "#dc2626"),
        "spillage":            ("IMDG", "#0369a1"),
        "first_aid":           ("MFAG", "#047857"),
        "general":             ("IMDG", "#475569"),
    }
    badge_code, badge_color = sop_badges.get(incident_type, ("IMDG", "#475569"))
    st.markdown(
        f'<div style="margin-top:-8px; margin-bottom:8px;">'
        f'<span style="background:{badge_color}; color:#fff; font-size:0.72rem;'
        f'font-weight:700; padding:3px 10px; border-radius:4px;'
        f'letter-spacing:1px;">WHL SOP {badge_code}</span>'
        f'&nbsp;<span style="font-size:0.78rem; color:#64748b;">'
        f'參考：{INCIDENT_SOP_MAP.get(incident_type, "IMDG Code")}</span>'
        f'</div>',
        unsafe_allow_html=True
    )

    additional = st.text_area(
        "額外情境說明（選填）",
        placeholder="例如：船艙內發現濃煙，疑似貨物起火，風速 15 節，船員 3 人在附近...",
        height=100
    )

    analyze_btn = st.button("🤖 開始 AI 分析", type="primary", use_container_width=True)

    if analyze_btn and un_input:
        st.session_state.last_un = un_input.strip()

        data = query_ems(un_input.strip())
        if data["found"]:
            col_a, col_b, col_c = st.columns(3)
            col_a.metric("UN 號碼",  data["un_number"])
            col_b.metric("物質名稱", data["proper_shipping_name"])
            col_c.metric("危險品類別", f"Class {data['hazard_class']}")

        st.markdown("---")
        st.markdown("#### 🤖 AI 應急建議")

        with st.spinner("AI 正在分析事故情境..."):
            result = analyze_incident(
                un_number     = un_input.strip(),
                incident_type = incident_type,
                additional_info = additional
            )

        st.markdown('<div class="ai-response-wrapper">', unsafe_allow_html=True)
        st.markdown(result)
        st.markdown('</div>', unsafe_allow_html=True)

    elif analyze_btn and not un_input:
        st.warning("⚠️ 請輸入 UN 號碼")


# ══════════════════════════════════════════════════════════════
# 頁面 3：積載隔離檢查
# ══════════════════════════════════════════════════════════════
elif page == "🔄 積載隔離檢查":

    st.markdown('<div class="main-title">🔄 積載隔離檢查</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">輸入 UN 號碼與貨櫃位置，檢查是否違反 IMDG 隔離規定</div>', unsafe_allow_html=True)

    with st.expander("📖 貨櫃位置格式說明 (BBRRTT)"):
        st.markdown("""
        | 欄位 | 說明 | 範例 |
        |------|------|------|
        | **BB** | Bay（貝位，船身前後位置） | 01, 03, 05… |
        | **RR** | Row（列位，左右位置） | 00=中心線, 01=左舷第1列, 02=右舷第1列 |
        | **TT** | Tier（層位，上下位置） | 02=艙內最底層, 82=甲板上第1層 |

        **完整範例：** `030282` = Bay 03, Row 02（右舷第1列）, Tier 82（甲板上第1層）

        > Tier 02–08 為艙內，82 以上為甲板上（每層遞增 2）
        """)

    st.markdown("---")
    st.markdown("#### 📦 貨物清單")
    st.caption("最多可新增 10 筆貨物，系統將逐一比對所有組合")

    col_add, col_remove, _ = st.columns([1, 1, 2])
    with col_add:
        if st.button("➕ 新增貨物", use_container_width=True):
            if len(st.session_state.cargo_list) < 10:
                n = len(st.session_state.cargo_list) + 1
                st.session_state.cargo_list.append(
                    {"un": "", "position": "", "label": f"貨物 {n}"}
                )
            else:
                st.warning("最多 10 筆貨物")
    with col_remove:
        if st.button("➖ 移除最後一筆", use_container_width=True):
            if len(st.session_state.cargo_list) > 2:
                st.session_state.cargo_list.pop()

    st.markdown("")

    validated_cargos = []

    for i, cargo in enumerate(st.session_state.cargo_list):
        col_label, col_un, col_pos, col_status = st.columns([1, 2, 2, 3])

        with col_label:
            st.markdown(f"<br><b>{cargo['label']}</b>", unsafe_allow_html=True)

        with col_un:
            un_val = st.text_input(
                "UN 號碼",
                value=cargo["un"],
                placeholder="例：1203",
                key=f"seg_un_{i}",
                label_visibility="collapsed" if i > 0 else "visible"
            )
            st.session_state.cargo_list[i]["un"] = un_val

        with col_pos:
            pos_val = st.text_input(
                "位置 (BBRRTT)",
                value=cargo["position"],
                placeholder="例：030282",
                key=f"seg_pos_{i}",
                label_visibility="collapsed" if i > 0 else "visible",
                max_chars=6
            )
            st.session_state.cargo_list[i]["position"] = pos_val

        with col_status:
            if un_val and pos_val:
                dg_data   = query_ems(un_val.strip())
                pos_valid = _validate_position(pos_val.strip())

                if dg_data["found"] and pos_valid:
                    st.success(
                        f"✅ {dg_data['proper_shipping_name'][:25]}… "
                        f"| Class {dg_data['hazard_class']} "
                        f"| {_format_position(pos_val.strip())}"
                    )
                    validated_cargos.append({
                        "label":    cargo["label"],
                        "un":       un_val.strip(),
                        "position": pos_val.strip(),
                        "data":     dg_data,
                    })
                elif not dg_data["found"]:
                    st.error(f"❌ UN{un_val} 查無資料")
                elif not pos_valid:
                    st.warning("⚠️ 位置格式錯誤，請輸入6位數字")
            elif un_val or pos_val:
                st.caption("請同時填入 UN 號碼與位置")

    st.markdown("---")

    if len(validated_cargos) >= 2:
        st.markdown("#### 🗺️ 貨物位置示意")
        _render_position_map(validated_cargos)

    check_btn = st.button(
        "🔄 執行隔離檢查",
        type="primary",
        use_container_width=True,
        disabled=len(validated_cargos) < 2
    )

    if len(validated_cargos) < 2:
        st.caption("⚠️ 請至少填入 2 筆有效貨物資料才能執行檢查")

    if check_btn and len(validated_cargos) >= 2:
        st.markdown("---")
        st.markdown("#### 📊 隔離檢查結果")

        pairs           = list(combinations(validated_cargos, 2))
        all_results     = []
        violation_count = 0

        for cargo_a, cargo_b in pairs:
            with st.spinner(f"檢查 {cargo_a['label']} × {cargo_b['label']}..."):
                result = check_segregation(
                    un_a  = cargo_a["un"],
                    un_b  = cargo_b["un"],
                    pos_a = cargo_a["position"],
                    pos_b = cargo_b["position"],
                )
                all_results.append({
                    "label_a": cargo_a["label"],
                    "label_b": cargo_b["label"],
                    "un_a":    cargo_a["un"],
                    "un_b":    cargo_b["un"],
                    "pos_a":   cargo_a["position"],
                    "pos_b":   cargo_b["position"],
                    "result":  result,
                })
                if any(kw in result for kw in ["違規", "違反", "不得", "禁止", "VIOLATION", "❌"]):
                    violation_count += 1

        if violation_count == 0:
            st.success(f"✅ 共檢查 {len(pairs)} 組配對，**未發現隔離違規**")
        else:
            st.error(f"🚨 共檢查 {len(pairs)} 組配對，發現 **{violation_count} 項隔離違規**，請立即處理！")

        for res in all_results:
            dist         = _calc_distance(res["pos_a"], res["pos_b"])
            is_violation = any(
                kw in res["result"]
                for kw in ["違規", "違反", "不得", "禁止", "VIOLATION", "❌"]
            )
            with st.expander(
                f"{'🚨' if is_violation else '✅'} "
                f"{res['label_a']} (UN{res['un_a']} @ {res['pos_a']})  ×  "
                f"{res['label_b']} (UN{res['un_b']} @ {res['pos_b']})  "
                f"｜距離約 {dist}",
                expanded=is_violation
            ):
                st.markdown('<div class="ai-response-wrapper">', unsafe_allow_html=True)
                st.markdown(res["result"])
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("---")
        st.download_button(
            label="📥 下載完整隔離檢查報告",
            data=_generate_segregation_report(validated_cargos, all_results, violation_count),
            file_name=f"segregation_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
            mime="text/plain",
            use_container_width=True
        )


# ══════════════════════════════════════════════════════════════
# 頁面 4：DG Bay Plan
# ══════════════════════════════════════════════════════════════
elif page == "🗺️ DG Bay Plan":

    st.markdown('<div class="main-title">🗺️ DG Bay Plan</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-title">上傳 DG 艙單，自動產生滅火介質視覺化積載圖</div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div class="warning-banner">'
        '⚠️ 本功能僅供船端緊急參考，實際操作請依官方 IMDG Code 及船長判斷'
        '</div>',
        unsafe_allow_html=True
    )

    # ── 色標說明 ────────────────────────────────────────────
    with st.expander("🎨 色標說明", expanded=False):
        legend = get_color_legend()
        cols   = st.columns(len(legend))
        for i, item in enumerate(legend):
            cols[i].markdown(
                f'<div style="background:{item["color_hex"]}; '
                f'color:white; padding:12px; border-radius:8px; '
                f'text-align:center; font-weight:700; min-height:100px;">'
                f'{item["label"]}<br>'
                f'<span style="font-size:0.75rem; font-weight:400;">'
                f'{item["media"]}</span><br>'
                f'<span style="font-size:0.7rem; opacity:0.85;">'
                f'EMS: {item["example"]}</span>'
                f'</div>',
                unsafe_allow_html=True
            )

    st.markdown("---")

    # ── 範本下載 + 檔案上傳 ─────────────────────────────────
    col_dl, col_up = st.columns([1, 2])

    with col_dl:
        st.markdown("#### 📥 下載範例模板")
        st.caption("不確定格式？先下載範例 Excel 填寫後上傳")
        st.download_button(
            label="⬇️ 下載 DG Manifest 範例",
            data=generate_sample_template(),
            file_name="dg_manifest_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_sample_template"
        )

    with col_up:
        st.markdown("#### 📤 上傳 DG 艙單")
        st.caption("支援 Excel (.xlsx)、CSV (.csv) 或船舶積載計畫 (.ASC)")
        uploaded_file = st.file_uploader(
            "支援 Excel (.xlsx)、CSV (.csv) 或 ASC 積載計畫",
            type=["xlsx", "xls", "csv", "asc"],
            label_visibility="collapsed",
            key="dg_bay_plan_uploader"
        )

    # ── 解析艙單（有上傳才執行）────────────────────────────
    if uploaded_file:
        file_bytes = uploaded_file.read()
        file_ext   = uploaded_file.name.split(".")[-1].lower()

        with st.spinner("解析艙單中..."):
            if file_ext in ("xlsx", "xls"):
                parsed_list, parse_warnings = parse_manifest_excel(file_bytes)
            elif file_ext == "csv":
                parsed_list, parse_warnings = parse_manifest_csv(file_bytes)
            elif file_ext == "asc":
                parsed_list, parse_warnings = parse_asc_file(file_bytes)
            else:
                parsed_list, parse_warnings = [], [f"❌ 不支援的檔案格式：.{file_ext}"]

        success_msgs = [w for w in parse_warnings if w.startswith("✅")]
        error_msgs   = [w for w in parse_warnings if w.startswith("❌")]
        warn_msgs    = [w for w in parse_warnings if w.startswith("⚠️")]

        for w in success_msgs:
            st.success(w)
        for w in error_msgs:
            st.error(w)

        if warn_msgs:
            if len(warn_msgs) <= 5:
                for w in warn_msgs:
                    st.warning(w)
            else:
                st.warning(f"⚠️ 共有 {len(warn_msgs)} 條警告訊息")
                with st.expander(f"展開查看所有警告（{len(warn_msgs)} 條）"):
                    for w in warn_msgs:
                        st.warning(w)

        if not parsed_list:
            st.error("❌ 未能解析任何有效貨物資料，請確認檔案格式")
            st.stop()

        if parsed_list[0].get("source") == "ASC":
            ship = parsed_list[0].get("ship_name", "")
            voy  = parsed_list[0].get("voyage",    "")
            if ship or voy:
                st.info(f"🚢 **{ship}** | 航次：{voy}")

        st.session_state["dg_cargo_list"] = parsed_list
        st.session_state["dg_bay_plan"]   = build_bay_plan(parsed_list)

    # ── 從 Session State 取值 ────────────────────────────────
    cargo_list = st.session_state.get("dg_cargo_list", [])
    bay_plan   = st.session_state.get("dg_bay_plan",   {})

    if not cargo_list:
        st.info("📤 請上傳 DG 艙單以開始分析")
        st.stop()

    # ════════════════════════════════════════════════════════
    # 區塊 A：統計摘要
    # ════════════════════════════════════════════════════════
    summary = get_manifest_summary(cargo_list)

    st.markdown("---")
    st.markdown("#### 📊 艙單摘要")

    if cargo_list[0].get("source") == "ASC":
        ship = cargo_list[0].get("ship_name", "")
        voy  = cargo_list[0].get("voyage",    "")
        if ship or voy:
            st.caption(f"🚢 {ship}　|　航次：{voy}")

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("🚢 全船裝載 DG 總數",          summary["total"])
    m2.metric("🟢 可用皮龍水柱滅火 (EMS F-A)",    summary["by_color"]["green"])
    m3.metric("🟡 使用滅火器滅火(CO2 / 乾粉) (EMS F-B~E)",  summary["by_color"]["yellow"])
    m4.metric("🔴 高危險碰水會產生化學反應 (EMS F-G+)", summary["by_color"]["red"])
    m5.metric("⚫ EMS 資料不完全",               summary["by_color"]["grey"])

    if summary["no_position"] > 0:
        st.warning(
            f"⚠️ 有 {summary['no_position']} 筆貨物缺少位置資料，"
            f"將不會顯示在 Bay Plan 圖上"
        )

    # ════════════════════════════════════════════════════════
    # 區塊 A2：危險櫃清單（新增）
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 📋 本航次危險品貨櫃清單")
    st.caption("依風險等級排序，所有危險品均積載於甲板（On Deck）")

    # 顏色對應的 emoji 標籤
    _color_emoji = {
        "red":    "🔴",
        "yellow": "🟡",
        "green":  "🟢",
        "grey":   "⚫",
    }

    # 依風險等級排序：red > yellow > green > grey，同色再依位置排序
    _risk_order = {"red": 0, "yellow": 1, "green": 2, "grey": 3}
    sorted_cargo = sorted(
        cargo_list,
        key=lambda c: (_risk_order.get(c["fire_color"], 3), c["position"])
    )

    dg_table_df = pd.DataFrame([{
        "風險":      _color_emoji.get(c["fire_color"], "⚫") + " " + c["fire_label"],
        "貨櫃號碼":  c["container_no"],
        "位置":      c["position"],
        "UN No":     f"UN{c['un_number']}",
        "Class":     c["hazard_class"],
        "PG":        c["packing_group"],
        "品名":      c["description"][:35] if c["description"] else "—",
        "Fire EMS":  c["fire_ems"]  if c["fire_ems"]  else "—",
        "Spill EMS": c["spill_ems"] if c["spill_ems"] else "—",
    } for c in sorted_cargo])

    # 依滅火分類上色
    def _highlight_dg_table(row):
        label = row["風險"]
        if "🔴" in label:
            return ["background-color:#fee2e2; color:#991b1b"] * len(row)
        elif "🟡" in label:
            return ["background-color:#fef9c3; color:#854d0e"] * len(row)
        elif "🟢" in label:
            return ["background-color:#dcfce7; color:#166534"] * len(row)
        else:
            return ["background-color:#f3f4f6; color:#374151"] * len(row)

    if not dg_table_df.empty:
        st.dataframe(
            dg_table_df.style.apply(_highlight_dg_table, axis=1),
            use_container_width=True,
            height=min(400, 45 + len(dg_table_df) * 38),  # 動態高度，最高 400px
        )
    else:
        st.info("無貨物資料")

    # ════════════════════════════════════════════════════════
    # 區塊 B：Bay Plan 視覺化（只顯示甲板）
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 🗺️ Bay Plan 危險櫃裝載視覺化")
    st.caption("依 IMDG 規定，危險品均積載於甲板，本圖僅顯示甲板層")

    if not bay_plan:
        st.info("所有貨物均缺少位置資料，無法繪製 Bay Plan")
    else:
        bay_options = sorted(bay_plan.keys())

        chosen_bay = st.selectbox(
            "選擇 Bay",
            options=["全部"] + [f"Bay {b:02d}" for b in bay_options],
            key="bayplan_selector"
        )

        bays_to_show = (
            {int(chosen_bay.replace("Bay ", "")): bay_plan[int(chosen_bay.replace("Bay ", ""))]}
            if chosen_bay != "全部"
            else bay_plan
        )

        rendered_any = False

        for bay_num, bay_data in sorted(bays_to_show.items()):
            dims = get_bay_dimensions(bay_data)

            # ── 只取甲板層（on_deck），忽略 in_hold ──────────
            tiers_deck = dims.get("tiers_deck", [])
            rows       = dims.get("rows", [])

            if not tiers_deck or not rows:
                continue   # 此 Bay 無甲板 DG，跳過

            rendered_any = True
            st.markdown(f"##### 🚢 Bay {bay_num:02d} — 甲板 Deck")

            fig = go.Figure()

            for ti, tier in enumerate(reversed(tiers_deck)):
                for ri, row in enumerate(rows):
                    cargos    = bay_data["on_deck"].get((row, tier), [])
                    cell      = get_cell_display(cargos)
                    bg_color  = cell["color_hex"] if cell else "#1e293b"
                    text_lbl  = cell["label"]     if cell else ""
                    hover_txt = (
                        cell["tooltip"] if cell
                        else f"Row {row:02d} / Tier {tier:02d}（空）"
                    )

                    fig.add_shape(
                        type="rect",
                        x0=ri,       y0=ti,
                        x1=ri + 1.0, y1=ti + 1.0,
                        fillcolor=bg_color,
                        line=dict(color="#334155", width=1),
                    )
                    if text_lbl:
                        fig.add_annotation(
                            x=ri + 0.5, y=ti + 0.5,
                            text=text_lbl.replace("\n", "<br>"),
                            showarrow=False,
                            font=dict(size=7, color="white"),   # ← 從 8 改為 7
                            align="center",
                            bgcolor="rgba(0,0,0,0.25)",         # ← 新增：半透明背景讓文字更清晰
                            borderpad=2,
                        )

                    fig.add_trace(go.Scatter(
                        x=[ri + 0.5], y=[ti + 0.5],
                        mode="markers",
                        marker=dict(size=30, opacity=0),
                        hovertext=hover_txt,
                        hoverinfo="text",
                        showlegend=False,
                    ))

            fig.update_xaxes(
                tickvals=[i + 0.5 for i in range(len(rows))],
                ticktext=[get_row_label(r) for r in rows],
                showgrid=False, zeroline=False,
                tickfont=dict(size=9, color="white"),
            )
            fig.update_yaxes(
                tickvals=[i + 0.5 for i in range(len(tiers_deck))],
                ticktext=[get_tier_label(t) for t in reversed(tiers_deck)],
                showgrid=False, zeroline=False,
                tickfont=dict(size=9, color="white"),
            )
            fig.update_layout(
                title=dict(
                    text=f"Bay {bay_num:02d} — Deck",
                    font=dict(color="white", size=13)
                ),
                height=max(180, len(tiers_deck) * 80 + 80),
                margin=dict(l=70, r=20, t=40, b=50),
                paper_bgcolor="#0f172a",
                plot_bgcolor="#0f172a",
                font=dict(color="white"),
            )
            st.plotly_chart(fig, use_container_width=True)

        if not rendered_any:
            st.info("選擇的 Bay 範圍內無甲板 DG 貨物")

    # ════════════════════════════════════════════════════════
    # 區塊 C：三秒判斷卡（Action Card）
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### ⚡ 三秒判斷卡")
    st.caption("選擇貨物，即時查看緊急處置摘要")

    cargo_options = {
        f"{c['container_no']} | UN{c['un_number']} | {c['description'][:25]}": c
        for c in cargo_list
    }
    selected_label = st.selectbox(
        "選擇貨物",
        options=list(cargo_options.keys()),
        label_visibility="collapsed",
        key="action_card_select"
    )

    if selected_label:
        cargo = cargo_options[selected_label]

        color_func = {
            "green":  st.success,
            "yellow": st.warning,
            "red":    st.error,
            "grey":   st.info,
        }.get(cargo["fire_color"], st.info)

        col_c1, col_c2 = st.columns(2)
        with col_c1:
            color_func(
                f"**滅火介質：{cargo['fire_label']}**\n\n"
                f"{cargo['fire_media']}"
            )
            st.markdown(
                f'<div class="info-card-green">'
                f'<b>✅ 應執行</b><br>{cargo["fire_do"]}'
                f'</div>',
                unsafe_allow_html=True
            )
        with col_c2:
            st.markdown(
                f'<div class="info-card">'
                f'<b>⛔ 禁止事項</b><br>{cargo["fire_dont"]}'
                f'</div>',
                unsafe_allow_html=True
            )
            st.markdown(
                f'<div class="info-card-blue">'
                f'<b>⚠️ 後續風險</b><br>{cargo["fire_risk"]}'
                f'</div>',
                unsafe_allow_html=True
            )

        with st.expander("📄 查看完整 IMDG 應急程序"):
            ems_data = query_ems(cargo["un_number"])
            if ems_data["found"]:
                tab_f, tab_s, tab_fa = st.tabs(["🔥 火災", "💧 洩漏", "🏥 急救"])
                with tab_f:
                    st.info(ems_data["emergency_action"].get("fire",      "無資料"))
                with tab_s:
                    st.info(ems_data["emergency_action"].get("spillage",  "無資料"))
                with tab_fa:
                    st.info(ems_data["emergency_action"].get("first_aid", "無資料"))
            else:
                st.warning(f"UN{cargo['un_number']} 查無 IMDG 資料")

    # ════════════════════════════════════════════════════════
    # 區塊 D：篩選器 + 貨物清單
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 🔍 篩選貨物清單")

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filter_color = st.multiselect(
            "依滅火分類篩選",
            options=["🟢 可用皮龍水柱滅火", "🟡 使用滅火器滅火", "🔴 高危險碰水會產生化學反應", "⚫ EMS 資料不完全"],
            default=[],
            key="filter_color"
        )
    with col_f2:
        all_bays_str = sorted(set(
            c["position"][0:2]
            for c in cargo_list if len(c["position"]) == 6
        ))
        filter_bay = st.multiselect(
            "依 Bay 篩選",
            options=[f"Bay {b}" for b in all_bays_str],
            default=[],
            key="filter_bay"
        )
    with col_f3:
        filter_class = st.multiselect(
            "依危險品類別篩選",
            options=sorted(set(
                c["hazard_class"] for c in cargo_list if c["hazard_class"]
            )),
            default=[],
            key="filter_class"
        )

    color_label_map = {
        "🟢 可用皮龍水柱滅火":    "green",
        "🟡 使用滅火器滅火(CO2 / 乾粉)":  "yellow",
        "🔴 高危險碰水會產生化學反應": "red",
        "⚫ EMS 資料不完全":      "grey",
    }
    filtered = cargo_list
    if filter_color:
        allowed_colors = {color_label_map[f] for f in filter_color}
        filtered = [c for c in filtered if c["fire_color"] in allowed_colors]
    if filter_bay:
        allowed_bays = {b.replace("Bay ", "") for b in filter_bay}
        filtered = [c for c in filtered if c["position"][0:2] in allowed_bays]
    if filter_class:
        filtered = [c for c in filtered if c["hazard_class"] in filter_class]

    st.caption(f"顯示 {len(filtered)} / {len(cargo_list)} 筆貨物")

    preview_df = pd.DataFrame([{
        "貨櫃號碼":  c["container_no"],
        "UN":        c["un_number"],
        "品名":      c["description"][:30],
        "Class":     c["hazard_class"],
        "PG":        c["packing_group"],
        "位置":      c["position"],
        "Fire EMS":  c["fire_ems"],
        "滅火分類":  c["fire_label"],
        "Spill EMS": c["spill_ems"],
    } for c in filtered])

    def _highlight_row(row):
        color_map = {
            "可用皮龍水柱滅火":       "background-color: #dcfce7; color: #166534",
            "使用滅火器滅火(CO2 / 乾粉)":     "background-color: #fef9c3; color: #854d0e",
            "高危險碰水會產生化學反應水":      "background-color: #fee2e2; color: #991b1b",
            "資料不完全":         "background-color: #f3f4f6; color: #374151",
        }
        return [color_map.get(row["滅火分類"], "")] * len(row)

    if not preview_df.empty:
        st.dataframe(
            preview_df.style.apply(_highlight_row, axis=1),
            use_container_width=True,
            height=400
        )
    else:
        st.info("目前篩選條件下無貨物資料")

    # ════════════════════════════════════════════════════════
    # 區塊 E：AI 全船風險摘要
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 🤖 AI 全船 DG 風險摘要")

    if st.button("🤖 產生 AI 風險摘要", use_container_width=True, key="ai_risk_summary"):

        # ── 預先在 Python 端做好分析，再餵給 AI ──────────────

        red_cargos    = [c for c in cargo_list if c["fire_color"] == "red"]
        yellow_cargos = [c for c in cargo_list if c["fire_color"] == "yellow"]
        green_cargos  = [c for c in cargo_list if c["fire_color"] == "green"]

        # ① 位置解析輔助
        def _pos_to_readable(pos: str) -> str:
            """030272 → Bay 03 / Row 02 / Tier 72"""
            if len(pos) == 6 and pos.isdigit():
                return f"Bay {pos[0:2]} / Row {pos[2:4]} / Tier {pos[4:6]}"
            return pos

        # ② 每個 Bay 的貨物分組（含顏色分布）
        bay_groups: dict[str, list] = {}
        for c in cargo_list:
            pos = c.get("position", "")
            if len(pos) == 6 and pos.isdigit():
                bay_key = f"Bay {pos[0:2]}"
                bay_groups.setdefault(bay_key, []).append(c)

        # ③ 找出混合風險 Bay（同一 Bay 內同時有不同滅火介質需求）
        conflict_bays = []
        for bay_key, bay_cargos in sorted(bay_groups.items()):
            colors_in_bay = set(c["fire_color"] for c in bay_cargos)
            has_red    = "red"    in colors_in_bay
            has_yellow = "yellow" in colors_in_bay
            has_green  = "green"  in colors_in_bay

            conflicts = []
            if has_red and has_green:
                conflicts.append("禁水貨物 ＋ 可用水貨物（水可能加劇禁水貨物危險）")
            if has_red and has_yellow:
                conflicts.append("禁水貨物 ＋ 非水介質貨物（需確認滅火介質不互相干擾）")
            if has_yellow and has_green:
                conflicts.append("非水介質貨物 ＋ 可用水貨物（滅火介質選擇需謹慎）")

            if conflicts:
                conflict_bays.append({
                    "bay":      bay_key,
                    "cargos":   bay_cargos,
                    "conflicts": conflicts,
                })

        # ④ 建立高風險貨物詳細清單（含可讀位置）
        def _cargo_line(c: dict) -> str:
            pos_readable = _pos_to_readable(c.get("position", "——"))
            return (
                f"  • {c['container_no']} | UN{c['un_number']} "
                f"| Class {c['hazard_class']} "
                f"| EMS Fire: {c['fire_ems'] or '—'} "
                f"| {pos_readable}"
                f"\n    品名：{c['description'][:40] if c['description'] else '未知'}"
            )

        red_lines    = "\n".join(_cargo_line(c) for c in red_cargos)    or "  （無）"
        yellow_lines = "\n".join(_cargo_line(c) for c in yellow_cargos) or "  （無）"

        # ⑤ 建立混合風險 Bay 說明
        if conflict_bays:
            conflict_lines_list = []
            for cb in conflict_bays:
                bay_cargo_summary = "\n".join(
                    f"    - {c['container_no']} UN{c['un_number']} "
                    f"({c['fire_label']}) @ {_pos_to_readable(c.get('position',''))}"
                    for c in cb["cargos"]
                )
                conflict_lines_list.append(
                    f"  ▶ {cb['bay']}（{len(cb['cargos'])} 個 DG 貨物）\n"
                    f"    衝突類型：{'；'.join(cb['conflicts'])}\n"
                    f"{bay_cargo_summary}"
                )
            conflict_section = "\n\n".join(conflict_lines_list)
        else:
            conflict_section = "  本航次各 Bay 無混合風險衝突"

        # ⑥ 每個 Bay 的滅火介質需求摘要（供建議滅火資源用）
        bay_summary_lines = []
        for bay_key, bay_cargos in sorted(bay_groups.items()):
            color_count = {}
            for c in bay_cargos:
                color_count[c["fire_label"]] = color_count.get(c["fire_label"], 0) + 1
            color_str = "、".join(f"{lbl}×{cnt}" for lbl, cnt in color_count.items())
            bay_summary_lines.append(f"  {bay_key}：{len(bay_cargos)} 個 DG（{color_str}）")
        bay_summary_section = "\n".join(bay_summary_lines) or "  （無位置資料）"

        # ══════════════════════════════════════════════════
        # 組合最終 Prompt
        # ══════════════════════════════════════════════════
        prompt = f"""
你是 WHL（長榮海運）的 FRM 應急顧問，請根據以下本航次 DG 危險品積載資料，
提供一份**船長級別的應急風險摘要報告**。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【本航次 DG 統計】
  總計：{len(cargo_list)} 個危險品貨櫃
  🔴 禁水/高危（EMS F-G/H/J 等）：{len(red_cargos)} 個
  🟡 非水介質（EMS F-B/C/E 等）  ：{len(yellow_cargos)} 個
  🟢 可用水（EMS F-A）            ：{len(green_cargos)} 個

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【各 Bay 積載分布】
{bay_summary_section}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【🔴 禁水/高危貨物明細】（需特別注意，禁止用水滅火）
{red_lines}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【🟡 非水介質貨物明細】（需使用乾粉/CO2/泡沫等）
{yellow_lines}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【⚠️ 混合風險 Bay 分析】（Python 預算結果，請據此說明）
{conflict_section}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

請依照以下格式輸出報告（繁體中文，重點加粗）：

### 1. 🔴 高風險貨物重點提示
針對每個禁水/高危貨物，說明：
- 主要危險特性（反應水、自燃、毒性等）
- 萬一起火/洩漏的首要處置原則
- **明確標示 Bay 位**（例如：Bay 03 / Row 02 / Tier 72）

### 2. ⚠️ 混合風險 Bay 警示
根據上方「混合風險 Bay 分析」資料，說明：
- 哪些 Bay 存在滅火介質衝突
- 具體衝突原因與潛在危險
- 建議的優先處置順序

### 3. 🚒 建議預置滅火資源
根據各 Bay 積載分布，具體建議：
- 各 Bay 應預置的滅火器材種類與數量
- 特別注意事項（例如：Bay XX 禁止預置水霧）

### 4. 📋 船長應知重點（不超過 5 條）
最關鍵的安全注意事項，每條需包含**具體 Bay 位或貨物資訊**

請確保每個 Bay 位都用「Bay XX」格式清楚標示，不要使用模糊的「某 Bay」描述。
"""

        with st.spinner("AI 正在分析全船風險..."):
            result = get_llm_response(
                system_prompt = SYSTEM_PROMPT,
                user_message  = prompt,
                max_tokens    = 2000,        # ← 從 1500 增加到 2000
                temperature   = 0.1,         # ← 從 0.2 降低到 0.1，讓輸出更精確
            )

        st.markdown('<div class="ai-response-wrapper">', unsafe_allow_html=True)
        st.markdown(result)
        st.markdown('</div>', unsafe_allow_html=True)


    # ════════════════════════════════════════════════════════
    # 區塊 F：匯出報告
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 📥 匯出報告")

    col_exp1, col_exp2 = st.columns(2)

    with col_exp1:
        if st.button("📊 產生 Excel 報告", use_container_width=True, key="gen_excel"):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

                wb = Workbook()
                ws = wb.active
                ws.title = "DG Bay Plan Report"

                headers = [
                    "貨櫃號碼", "UN", "品名", "Class", "PG",
                    "位置", "Fire EMS", "滅火分類", "滅火介質", "Spill EMS"
                ]
                header_fill = PatternFill("solid", fgColor="1e293b")
                thin_border = Border(
                    left=Side(style="thin"),  right=Side(style="thin"),
                    top=Side(style="thin"),   bottom=Side(style="thin"),
                )

                for col_idx, header in enumerate(headers, 1):
                    cell           = ws.cell(row=1, column=col_idx, value=header)
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border    = thin_border

                excel_colors = {
                    "green":  "dcfce7",
                    "yellow": "fef9c3",
                    "red":    "fee2e2",
                    "grey":   "f3f4f6",
                }

                sorted_cargos = sorted(
                    cargo_list,
                    key=lambda x: (
                        {"red": 0, "yellow": 1, "green": 2, "grey": 3}.get(x["fire_color"], 3),
                        x["position"]
                    )
                )

                for row_idx, c in enumerate(sorted_cargos, 2):
                    values = [
                        c["container_no"], c["un_number"],     c["description"],
                        c["hazard_class"], c["packing_group"], c["position"],
                        c["fire_ems"],     c["fire_label"],    c["fire_media"],
                        c["spill_ems"],
                    ]
                    fill = PatternFill(
                        "solid",
                        fgColor=excel_colors.get(c["fire_color"], "f3f4f6")
                    )
                    for col_idx, val in enumerate(values, 1):
                        cell           = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.fill      = fill
                        cell.border    = thin_border
                        cell.alignment = Alignment(horizontal="left")

                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

                ws.freeze_panes = "A2"

                buf = io.BytesIO()
                wb.save(buf)

                st.download_button(
                    label="⬇️ 下載 Excel 報告",
                    data=buf.getvalue(),
                    file_name=f"dg_bayplan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_excel_report"
                )
            except ImportError:
                st.error("請安裝 openpyxl：pip install openpyxl")

    with col_exp2:
        # 船名航次資訊（用於報告標頭）
        _ship = cargo_list[0].get("ship_name", "") if cargo_list else ""
        _voy  = cargo_list[0].get("voyage",    "") if cargo_list else ""
        _ship_line = f"  船名：{_ship}　航次：{_voy}" if _ship or _voy else ""

        lines = [
            "=" * 65,
            "  DG CARGO GUARDIAN — Bay Plan 滅火介質報告",
            f"  產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ]
        if _ship_line:
            lines.append(_ship_line)
        lines += [
            "=" * 65, "",
            f"  總貨物數    ：{summary['total']}",
            f"  🟢 可用水    ：{summary['by_color']['green']}",
            f"  🟡 非水介質  ：{summary['by_color']['yellow']}",
            f"  🔴 禁水/高危 ：{summary['by_color']['red']}",
            f"  ⚫ 未知      ：{summary['by_color']['grey']}",
            "", "-" * 65,
            "  【依風險等級 + 位置排序】",
            "-" * 65,
        ]
        for color_key, color_label in [
            ("red",    "🔴 禁水/高危"),
            ("yellow", "🟡 非水介質"),
            ("green",  "🟢 可用水"),
            ("grey",   "⚫ 未知"),
        ]:
            group = sorted(
                [c for c in cargo_list if c["fire_color"] == color_key],
                key=lambda c: c["position"]
            )
            if not group:
                continue
            lines.append(f"\n  {color_label}（{len(group)} 個）")
            lines.append("  " + "-" * 60)
            for c in group:
                lines.append(
                    f"  {c['container_no']} | UN{c['un_number']:4s} | "
                    f"{c['description'][:28]:28s} | "
                    f"EMS {c['fire_ems']:4s} | 位置 {c['position']}"
                )
        lines += [
            "", "=" * 65,
            "  ⚠️  本報告僅供參考，實際操作請依 IMDG Code 官方規定",
            "=" * 65,
        ]

        st.download_button(
            label="📄 下載文字報告",
            data="\n".join(lines),
            file_name=f"dg_bayplan_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
            mime="text/plain",
            use_container_width=True,
            key="dl_text_report"
        )


# ══════════════════════════════════════════════════════════════
# 頁面 5：自由問答
# ══════════════════════════════════════════════════════════════
elif page == "💬 自由問答":

    st.markdown('<div class="main-title">💬 自由問答</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">詢問任何 IMDG / 危險品相關問題</div>', unsafe_allow_html=True)

    # 聊天記錄顯示
    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    # 選填 UN 號碼
    with st.expander("🔖 附加 UN 號碼資料（選填）"):
        ref_un = st.text_input(
            "UN 號碼",
            placeholder="填入後 AI 會參考該危險品資料回答",
            key="chat_un"
        )

    st.markdown(
        '<div style="font-size:0.72rem; color:#475569; text-transform:uppercase;'
        'letter-spacing:1.5px; margin:10px 0 6px 0;">📖 WHL SOP 快捷問題</div>',
        unsafe_allow_html=True
    )

    quick_questions = {
        "🔥 甲板貨櫃失火，CO2 釋放前需確認哪些事項？":
            "依 WHL 3-3，甲板貨櫃失火時，CO2 釋放前需確認哪些事項？請逐條列出。",
        "💧 貨艙失火，何時應放棄探火直接釋放 CO2？":
            "依 WHL 3-4，貨艙貨櫃失火時，何種情況下應放棄探火，直接釋放 CO2？",
        "🔧 機艙失火，CO2 釋放後需密封多久？":
            "依 WHL 1-5，機艙失火釋放 CO2 後，機艙需保持密封多少小時？原因為何？",
        "📦 貨櫃落海，惡劣天候甲板作業需符合哪些條件？":
            "依 WHL 3-2，貨櫃落海後派員至甲板作業，需符合哪些安全條件？",
        "☎️ 緊急事故發生後，何時需通知海技部？":
            "WHL 各類緊急事故（失火/洩漏/落海），通知海技部（Maritech Division）的時機與內容要點為何？",
    }

    cols = st.columns(2)
    for i, (btn_label, question_text) in enumerate(quick_questions.items()):
        with cols[i % 2]:
            if st.button(btn_label, use_container_width=True, key=f"quick_q_{i}"):
                st.session_state.chat_history.append({
                    "role": "user", "content": question_text
                })
                st.rerun()

    # 輸入框
    user_input = st.chat_input("輸入你的問題，例如：Class 3 危險品的主要火災風險是什麼？")

    if user_input:
        st.session_state.chat_history.append({
            "role": "user",
            "content": user_input
        })
        with st.chat_message("user"):
            st.markdown(user_input)

        with st.chat_message("assistant"):
            with st.spinner("思考中..."):
                response = ask_dg_question(
                    question  = user_input,
                    un_number = ref_un.strip() if ref_un else None
                )
            st.markdown(response)

        st.session_state.chat_history.append({
            "role":    "assistant",
            "content": response
        })

    # 清除對話
    if st.session_state.chat_history:
        if st.button("🗑️ 清除對話記錄"):
            st.session_state.chat_history = []
            st.rerun()

